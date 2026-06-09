import os
import tempfile
import unittest

from openpyxl import Workbook, load_workbook

from term_extractor_app.diff_excel import (
    CompareMeta,
    apply_highlight_to_records,
    apply_highlight_from_cache,
    compare_excel_files,
    compare_paths,
    export_cached_diff_records,
    export_diff_records,
    read_cached_diff_preview,
    run_compare_to_cache,
)


def create_sample_workbook(path: str, value_b2: str, value_c3: str) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Main"
    worksheet["A1"] = "ID"
    worksheet["B1"] = "Text"
    worksheet["C1"] = "Comment"
    worksheet["A2"] = 1
    worksheet["B2"] = value_b2
    worksheet["C2"] = "same"
    worksheet["A3"] = 2
    worksheet["B3"] = "same"
    worksheet["C3"] = value_c3
    workbook.save(path)
    workbook.close()


class DiffExcelTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.path_a = os.path.join(self.temp_dir.name, "a.xlsx")
        self.path_b = os.path.join(self.temp_dir.name, "b.xlsx")
        create_sample_workbook(self.path_a, "Alpha", "Old note")
        create_sample_workbook(self.path_b, "Beta", "New note")

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    def test_compare_excel_files_returns_expected_cells(self) -> None:
        diffs = compare_excel_files(self.path_a, self.path_b)
        addresses = sorted(diff.cell_address for diff in diffs)
        self.assertEqual(addresses, ["B2", "C3"])

    def test_compare_paths_directory_mode_matches_same_name_files(self) -> None:
        dir_a = os.path.join(self.temp_dir.name, "dir_a")
        dir_b = os.path.join(self.temp_dir.name, "dir_b")
        os.makedirs(dir_a, exist_ok=True)
        os.makedirs(dir_b, exist_ok=True)
        file_a = os.path.join(dir_a, "same.xlsx")
        file_b = os.path.join(dir_b, "same.xlsx")
        create_sample_workbook(file_a, "One", "Same")
        create_sample_workbook(file_b, "Two", "Same")

        diffs, meta = compare_paths(dir_a, dir_b)

        self.assertEqual(len(diffs), 1)
        self.assertEqual(diffs[0].cell_address, "B2")
        self.assertEqual(meta.mode_label, "目录对目录")
        self.assertEqual(meta.matched_pairs, 1)

    def test_apply_highlight_to_records_marks_target_workbook(self) -> None:
        diffs = compare_excel_files(self.path_a, self.path_b)
        changed_cells, workbook_count = apply_highlight_to_records(diffs, "A", "#FFD966")

        self.assertEqual(changed_cells, 2)
        self.assertEqual(workbook_count, 1)

        workbook = load_workbook(self.path_a)
        try:
            worksheet = workbook["Main"]
            self.assertTrue((worksheet["B2"].fill.start_color.rgb or "").endswith("FFD966"))
            self.assertTrue((worksheet["C3"].fill.start_color.rgb or "").endswith("FFD966"))
        finally:
            workbook.close()

    def test_export_diff_records_writes_result_file(self) -> None:
        diffs = compare_excel_files(self.path_a, self.path_b)
        output_path = os.path.join(self.temp_dir.name, "diff_output.xlsx")

        result = export_diff_records(diffs, output_path)

        self.assertEqual(result["diff_count"], 2)
        self.assertTrue(os.path.exists(output_path))

        workbook = load_workbook(output_path)
        try:
            worksheet = workbook["差异列表"]
            self.assertEqual(worksheet["A2"].value, "a.xlsx")
            self.assertEqual(worksheet["D2"].value, "B2")
            self.assertEqual(worksheet["E2"].value, "Alpha")
            self.assertEqual(worksheet["F2"].value, "Beta")
        finally:
            workbook.close()

    def test_run_compare_to_cache_limits_preview_and_preserves_total(self) -> None:
        result = run_compare_to_cache(self.path_a, self.path_b, preview_limit=1)

        self.assertEqual(result["total_count"], 2)
        self.assertEqual(len(result["preview_records"]), 1)
        self.assertTrue(result["preview_truncated"])
        self.assertTrue(os.path.exists(result["cache_file"]))

        preview = read_cached_diff_preview(result["cache_file"], limit=5)
        self.assertEqual(preview["matched_count"], 2)
        self.assertEqual(len(preview["records"]), 2)

    def test_export_and_highlight_from_cache_work_with_query(self) -> None:
        result = run_compare_to_cache(self.path_a, self.path_b, preview_limit=10)
        output_path = os.path.join(self.temp_dir.name, "diff_filtered.xlsx")

        exported = export_cached_diff_records(result["cache_file"], output_path, query="old note")
        self.assertEqual(exported["diff_count"], 1)
        self.assertTrue(os.path.exists(output_path))

        changed_cells, workbook_count = apply_highlight_from_cache(result["cache_file"], "A", "#9DC3E6", query="old note")
        self.assertEqual(changed_cells, 1)
        self.assertEqual(workbook_count, 1)

        workbook = load_workbook(self.path_a)
        try:
            worksheet = workbook["Main"]
            self.assertTrue((worksheet["C3"].fill.start_color.rgb or "").endswith("9DC3E6"))
            self.assertFalse((worksheet["B2"].fill.start_color.rgb or "").endswith("9DC3E6"))
        finally:
            workbook.close()


if __name__ == "__main__":
    unittest.main()
