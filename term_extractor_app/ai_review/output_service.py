from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .config import OUTPUTS_DIR, ensure_directories
from .database import get_connection


NORMAL_HEADERS = [
    "来源文件",
    "sheet / segment ID",
    "原始行号",
    "原文",
    "译文",
    "是否有问题",
    "问题类型",
    "问题说明",
    "修改建议",
]


def generate_review_excel(task_id: str) -> Path:
    ensure_directories()
    task = _fetch_task(task_id)
    config = task["config"]
    rows = _fetch_rows(task_id)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = OUTPUTS_DIR / f"review_result_{timestamp}.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "审校结果"
    enable_forbidden = bool(config.get("enable_forbidden_check"))
    if config.get("mode") == "forbidden_only":
        review_type_keys = []
        headers = ["来源文件", "sheet / segment ID", "原始行号", "原文", "译文", "禁用词检查情况"]
    elif config.get("mode") == "directional":
        review_type_keys = [item["key"] for item in config.get("review_types", [])]
        headers = ["来源文件", "sheet / segment ID", "原始行号", "原文", "译文", *review_type_keys]
    else:
        review_type_keys = []
        headers = NORMAL_HEADERS
    if enable_forbidden and config.get("mode") != "forbidden_only":
        headers.append("禁用词检查情况")
    sheet.append(headers)

    for row in rows:
        location = row["sheet_name"] or row["segment_id"] or ""
        forbidden = row["matched_words"] or ""
        if config.get("mode") == "forbidden_only":
            sheet.append(
                [
                    row["source_file"],
                    location,
                    row["row_number"] or "",
                    row["source_text"] or "",
                    row["target_text"] or "",
                    forbidden,
                ]
            )
            continue

        if config.get("mode") == "directional":
            checks = _loads_checks(row["directional_checks_json"])
            values = [
                row["source_file"],
                location,
                row["row_number"] or "",
                row["source_text"] or "",
                row["target_text"] or "",
                *[row["error_message"] if row["status"] == "failed" else checks.get(key, "") for key in review_type_keys],
            ]
            if enable_forbidden:
                values.append(forbidden)
            sheet.append(values)
            continue

        has_issue = ""
        if row["status"] == "failed":
            has_issue = "失败"
        elif row["has_issue"] is not None:
            has_issue = "是" if row["has_issue"] else "否"
        issue = row["error_message"] or row["issue"] or ""
        values = [
            row["source_file"],
            location,
            row["row_number"] or "",
            row["source_text"] or "",
            row["target_text"] or "",
            has_issue,
            row["issue_type"] or "",
            issue,
            row["suggestion"] or "",
        ]
        if enable_forbidden:
            values.append(forbidden)
        sheet.append(values)

    _format_sheet(sheet)
    workbook.save(output_path)
    return output_path


def _fetch_rows(task_id: str) -> list[Any]:
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT r.status, r.has_issue, r.issue_type, r.issue, r.suggestion,
                   r.directional_checks_json, r.error_message,
                   COALESCE(f.matched_words, '') AS matched_words,
                   i.source_file, i.sheet_name, i.segment_id, i.row_number,
                   i.source_text, i.target_text
            FROM review_results r
            JOIN file_items i ON i.id = r.item_id
            LEFT JOIN forbidden_results f ON f.task_id = r.task_id AND f.item_id = r.item_id
            WHERE r.task_id = ?
            ORDER BY i.item_order ASC
            """,
            (task_id,),
        ).fetchall()
    return rows


def _fetch_task(task_id: str) -> dict[str, Any]:
    from .database import loads_json

    with get_connection() as conn:
        row = conn.execute("SELECT config_json FROM review_tasks WHERE id = ?", (task_id,)).fetchone()
    if not row:
        return {"config": {}}
    return {"config": loads_json(row["config_json"], {})}


def _loads_checks(text: str) -> dict[str, str]:
    from .database import loads_json

    data = loads_json(text, {})
    return data if isinstance(data, dict) else {}


def _format_sheet(sheet: Any) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="EAF2EF")
    header_font = Font(bold=True, color="1D2428")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    base_widths = [22, 18, 10, 42, 42]
    widths = [*base_widths, *([24] * max(0, sheet.max_column - len(base_widths)))]
    for index, width in enumerate(widths[: sheet.max_column], start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = width

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
