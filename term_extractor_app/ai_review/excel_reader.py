from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _sheet_columns(sheet: Any) -> list[dict[str, Any]]:
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    headers = [_cell_to_text(value) for value in header_row] if header_row else []
    used_indexes: set[int] = {index for index, header in enumerate(headers) if header}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        for index, value in enumerate(row):
            if _cell_to_text(value):
                used_indexes.add(index)

    return [
        {
            "index": index,
            "letter": get_column_letter(index + 1),
            "header": headers[index] if index < len(headers) else "",
        }
        for index in sorted(used_indexes)
    ]


def read_excel_headers(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    headers_by_sheet: dict[str, list[str]] = {}
    columns_by_sheet: dict[str, list[dict[str, Any]]] = {}
    ordered_headers: list[str] = []
    seen: set[str] = set()

    for sheet in workbook.worksheets:
        columns = _sheet_columns(sheet)
        columns_by_sheet[sheet.title] = columns
        headers = [column["header"] for column in columns if column["header"]]
        headers_by_sheet[sheet.title] = headers
        for header in headers:
            if header not in seen:
                seen.add(header)
                ordered_headers.append(header)

    workbook.close()
    if not any(columns_by_sheet.values()):
        raise ValueError("没有读取到可用列，请确认 Excel 中存在非空列")

    return {
        "headers": ordered_headers,
        "headers_by_sheet": headers_by_sheet,
        "columns_by_sheet": columns_by_sheet,
        "sheet_names": list(headers_by_sheet.keys()),
    }


def read_excel_items(path: Path, source_column: str, target_column: str, source_file: str) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    items: list[dict[str, Any]] = []
    order = 0

    for sheet in workbook.worksheets:
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        headers = [_cell_to_text(value) for value in header_row] if header_row else []
        if source_column not in headers or target_column not in headers:
            continue

        source_index = headers.index(source_column)
        target_index = headers.index(target_column)

        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            source_text = _cell_to_text(row[source_index]) if source_index < len(row) else ""
            target_text = _cell_to_text(row[target_index]) if target_index < len(row) else ""
            if not source_text and not target_text:
                continue

            status_note = ""
            if not source_text:
                status_note = "原文为空"
            elif not target_text:
                status_note = "译文为空"

            order += 1
            items.append(
                {
                    "source_file": source_file,
                    "sheet_name": sheet.title,
                    "segment_id": None,
                    "row_number": row_number,
                    "source_text": source_text,
                    "target_text": target_text,
                    "status_note": status_note,
                    "item_order": order,
                    "info": [],
                    "source_column": get_column_letter(source_index + 1),
                    "target_column": get_column_letter(target_index + 1),
                }
            )

    workbook.close()
    if not items:
        raise ValueError("没有读取到可审校条目，请检查原文列和译文列是否选择正确")
    return items


def read_excel_items_by_mapping(path: Path, mapping: dict[str, Any], source_file: str) -> list[dict[str, Any]]:
    validate_excel_mapping(mapping)
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_map = {sheet.title: sheet for sheet in workbook.worksheets}
    items: list[dict[str, Any]] = []
    order = 0

    for sheet_config in mapping.get("sheets", []):
        sheet_name = str(sheet_config.get("sheet_name") or "")
        sheet = sheet_map.get(sheet_name)
        if sheet is None:
            continue

        mappings = sheet_config.get("mappings", [])
        if not isinstance(mappings, list):
            continue

        for column_mapping in mappings:
            source_index = _as_column_index(column_mapping.get("source_column"))
            target_index = _as_column_index(column_mapping.get("target_column"))
            if source_index is None or target_index is None:
                continue

            info_columns = _normalize_info_columns(column_mapping.get("info_columns", []))
            source_label = get_column_letter(source_index + 1)
            target_label = get_column_letter(target_index + 1)

            for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                source_text = _cell_to_text(row[source_index]) if source_index < len(row) else ""
                target_text = _cell_to_text(row[target_index]) if target_index < len(row) else ""
                if not source_text and not target_text:
                    continue

                info_items = []
                for info_column in info_columns:
                    index = info_column["column"]
                    value = _cell_to_text(row[index]) if index < len(row) else ""
                    if value:
                        info_items.append(
                            {
                                "category": info_column.get("category", ""),
                                "value": value,
                                "column": get_column_letter(index + 1),
                            }
                        )

                status_note = ""
                if not source_text:
                    status_note = "原文为空"
                elif not target_text:
                    status_note = "译文为空"

                order += 1
                items.append(
                    {
                        "source_file": source_file,
                        "sheet_name": sheet.title,
                        "segment_id": None,
                        "row_number": row_number,
                        "source_text": source_text,
                        "target_text": target_text,
                        "status_note": status_note,
                        "item_order": order,
                        "info": info_items,
                        "source_column": source_label,
                        "target_column": target_label,
                    }
                )

    workbook.close()
    if not items:
        raise ValueError("没有读取到可审校条目，请检查列映射是否正确")
    return items


def validate_excel_mapping(mapping: dict[str, Any]) -> None:
    total_mappings = 0
    for sheet_config in mapping.get("sheets", []):
        sheet_name = str(sheet_config.get("sheet_name") or "")
        mappings = sheet_config.get("mappings", [])
        if not sheet_name or not isinstance(mappings, list):
            continue
        source_columns: set[int] = set()
        target_columns: set[int] = set()
        for column_mapping in mappings:
            source_index = _as_column_index(column_mapping.get("source_column"))
            target_index = _as_column_index(column_mapping.get("target_column"))
            if source_index is None:
                raise ValueError(f"{sheet_name} 存在未选择原文列的配置")
            if target_index is None:
                raise ValueError(f"{sheet_name} 的 {get_column_letter(source_index + 1)} 列缺少译文列")
            if source_index in source_columns:
                raise ValueError(f"{sheet_name} 的 {get_column_letter(source_index + 1)} 列被重复设置为原文列")
            if target_index in target_columns:
                raise ValueError(f"{sheet_name} 的 {get_column_letter(target_index + 1)} 列被多个原文列共用为译文列")
            if source_index == target_index:
                raise ValueError(f"{sheet_name} 的原文列和译文列不能相同")
            for info_column in _normalize_info_columns(column_mapping.get("info_columns", [])):
                info_index = info_column["column"]
                if info_index == target_index:
                    raise ValueError(f"{sheet_name} 的 {get_column_letter(info_index + 1)} 列不能同时作为译文列和信息列")
                if info_index == source_index:
                    raise ValueError(f"{sheet_name} 的 {get_column_letter(info_index + 1)} 列不能同时作为原文列和信息列")
            source_columns.add(source_index)
            target_columns.add(target_index)
            total_mappings += 1

    if total_mappings == 0:
        raise ValueError("请至少配置一组原文列和译文列")


def _as_column_index(value: Any) -> int | None:
    try:
        index = int(value)
    except (TypeError, ValueError):
        return None
    return index if index >= 0 else None


def _normalize_info_columns(value: Any) -> list[dict[str, Any]]:
    if not isinstance(value, list):
        return []
    normalized = []
    for item in value:
        if not isinstance(item, dict):
            continue
        index = _as_column_index(item.get("column"))
        if index is not None:
            normalized.append({"column": index, "category": str(item.get("category") or "").strip()})
    return normalized
