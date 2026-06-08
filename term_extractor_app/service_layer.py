"""UI-neutral task facade for desktop UI and future WebUI."""

from __future__ import annotations

import asyncio
import logging
import threading
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable, Dict, List, Optional

from openpyxl import load_workbook

from .constants import FAILURE_SHEET, NONTRANS_REGEX_SHEET, REVIEW_SHEET, TERM_LIBRARY_SHEET
from .logging_utils import configure_file_logger, reset_file_logger
from .models import AppSettings, RuntimeTaskState, TaskInput
from .pipeline import TaskCancelledError, TermExtractionService
from .storage import AppPaths, RuntimeCacheStore, SettingsStore, get_app_paths
from .telemetry import infer_model_tier, track_event


@dataclass
class TaskSnapshot:
    is_running: bool
    can_resume: bool = False
    stage: str = "IDLE"
    stage_label: str = ""
    progress_current: int = 0
    progress_total: int = 0
    current_batch: int = 0
    total_batches: int = 0
    success_count: int = 0
    failure_count: int = 0
    retry_count: int = 0
    current_concurrency: int = 0
    output_file: str = ""
    last_error: str = ""
    message: str = ""
    logs: List[str] = field(default_factory=list)
    stats: Dict[str, object] = field(default_factory=dict)

    def to_dict(self) -> Dict[str, object]:
        return {
            "is_running": self.is_running,
            "can_resume": self.can_resume,
            "stage": self.stage,
            "stage_label": self.stage_label,
            "progress_current": self.progress_current,
            "progress_total": self.progress_total,
            "current_batch": self.current_batch,
            "total_batches": self.total_batches,
            "success_count": self.success_count,
            "failure_count": self.failure_count,
            "retry_count": self.retry_count,
            "current_concurrency": self.current_concurrency,
            "output_file": self.output_file,
            "last_error": self.last_error,
            "message": self.message,
            "logs": list(self.logs),
            "stats": dict(self.stats),
        }


@dataclass
class ResultSummary:
    output_file: str = ""
    exists: bool = False
    term_library_count: int = 0
    review_count: int = 0
    failure_count: int = 0
    nontrans_regex_count: int = 0
    error: str = ""

    def to_dict(self) -> Dict[str, object]:
        return {
            "output_file": self.output_file,
            "exists": self.exists,
            "term_library_count": self.term_library_count,
            "review_count": self.review_count,
            "failure_count": self.failure_count,
            "nontrans_regex_count": self.nontrans_regex_count,
            "error": self.error,
        }


class ExtractionTaskFacade:
    """Small service boundary that keeps UI code out of extraction logic."""

    def __init__(
        self,
        paths: Optional[AppPaths] = None,
        logger: Optional[logging.Logger] = None,
        settings_store: Optional[SettingsStore] = None,
        runtime_store: Optional[RuntimeCacheStore] = None,
        service_factory: Optional[Callable[..., TermExtractionService]] = None,
    ):
        self.paths = paths or get_app_paths()
        self.settings_store = settings_store or SettingsStore(self.paths)
        self.runtime_store = runtime_store or RuntimeCacheStore(self.paths)
        self.logger = logger or configure_file_logger(self.paths)
        self.service_factory = service_factory or TermExtractionService
        self._lock = threading.RLock()
        self._thread: Optional[threading.Thread] = None
        self._stop_event = threading.Event()
        self._logs: List[str] = []
        self._last_progress: Dict[str, object] = {}
        self._is_running = False
        self._output_file = ""
        self._last_error = ""
        self._active_telemetry_context: Dict[str, object] = {}

    def load_settings(self) -> AppSettings:
        return self.settings_store.load()

    def save_settings(self, settings: AppSettings) -> AppSettings:
        self.settings_store.save(settings)
        return settings

    def can_start(self) -> bool:
        with self._lock:
            return not self._is_running

    def can_resume(self) -> bool:
        with self._lock:
            return (not self._is_running) and self.runtime_store.exists()

    def start(self, task_input: Optional[TaskInput], resume: bool = False, settings: Optional[AppSettings] = None) -> None:
        with self._lock:
            if self._is_running:
                raise RuntimeError("\u5df2\u6709\u4efb\u52a1\u6b63\u5728\u8fd0\u884c\u3002")
            if resume and not self.runtime_store.exists():
                raise RuntimeError("\u6ca1\u6709\u53ef\u7ee7\u7eed\u7684\u4efb\u52a1\u3002")
            active_settings = settings or self.load_settings()
            self.logger = reset_file_logger(self.paths)
            self._logs = []
            self._last_progress = {}
            self._output_file = ""
            self._last_error = ""
            self._active_telemetry_context = _build_telemetry_context(
                settings=active_settings,
                task_input=task_input,
                runtime=self.runtime_store.load() if resume else None,
            )
            _track_text_preprocess_start(self._active_telemetry_context)
            self._stop_event.clear()
            self._is_running = True
            self._thread = threading.Thread(
                target=self._run_task,
                args=(active_settings, task_input, resume),
                name="extraction-task",
                daemon=True,
            )
            self._thread.start()

    def resume(self, settings: Optional[AppSettings] = None) -> None:
        self.start(None, resume=True, settings=settings)

    def stop(self) -> None:
        self._stop_event.set()

    def clear_runtime_cache(self) -> None:
        with self._lock:
            if self._is_running:
                raise RuntimeError("\u4efb\u52a1\u8fd0\u884c\u4e2d\uff0c\u4e0d\u80fd\u6e05\u7a7a\u7f13\u5b58\u3002")
            self.runtime_store.clear()
            self._logs.append("\u8fd0\u884c\u7f13\u5b58\u5df2\u6e05\u7a7a\u3002")
            self._last_progress = {}
            self._last_error = ""

    def wait(self, timeout: Optional[float] = None) -> bool:
        thread = self._thread
        if thread is None:
            return True
        thread.join(timeout)
        return not thread.is_alive()

    def snapshot(self) -> TaskSnapshot:
        with self._lock:
            runtime = self.runtime_store.load()
            stage = str(self._last_progress.get("stage") or (runtime.stage if runtime else "IDLE") or "IDLE")
            stats = dict(runtime.stats if runtime else {})
            progress_payload = dict(self._last_progress or {})
            progress_stats = {
                "progress_current": progress_payload.get("progress_current", stats.get("progress_current", 0)),
                "progress_total": progress_payload.get("progress_total", stats.get("progress_total", 0)),
                "current_batch": progress_payload.get("current_batch", stats.get("current_batch", 0)),
                "total_batches": progress_payload.get("total_batches", stats.get("total_batches", 0)),
                "success_count": progress_payload.get("success_count", stats.get("success_count", 0)),
                "failure_count": progress_payload.get("failure_count", stats.get("failure_count", 0)),
                "retry_count": progress_payload.get("retry_count", stats.get("retry_count", 0)),
                "current_concurrency": progress_payload.get(
                    "current_concurrency", stats.get("current_concurrency", 0)
                ),
            }
            visible_stats = {**stats, **progress_payload, **progress_stats}
            output_file = self._output_file or str(runtime.output_file if runtime else "")
            last_error = self._last_error or str(runtime.last_error if runtime else "")
            return TaskSnapshot(
                is_running=self._is_running,
                can_resume=(not self._is_running) and self.runtime_store.exists(),
                stage=stage,
                stage_label=str(self._last_progress.get("stage_label", "")),
                progress_current=int(self._last_progress.get("progress_current", stats.get("progress_current", 0)) or 0),
                progress_total=int(self._last_progress.get("progress_total", stats.get("progress_total", 0)) or 0),
                current_batch=int(self._last_progress.get("current_batch", stats.get("current_batch", 0)) or 0),
                total_batches=int(self._last_progress.get("total_batches", stats.get("total_batches", 0)) or 0),
                success_count=int(self._last_progress.get("success_count", stats.get("success_count", 0)) or 0),
                failure_count=int(self._last_progress.get("failure_count", stats.get("failure_count", 0)) or 0),
                retry_count=int(self._last_progress.get("retry_count", stats.get("retry_count", 0)) or 0),
                current_concurrency=int(
                    self._last_progress.get("current_concurrency", stats.get("current_concurrency", 0)) or 0
                ),
                output_file=output_file,
                last_error=last_error,
                message=str(self._last_progress.get("message", "")),
                logs=list(self._logs[-200:]),
                stats=visible_stats,
            )

    def result_summary(self, output_file: Optional[str] = None) -> ResultSummary:
        with self._lock:
            runtime = self.runtime_store.load()
            resolved_output = output_file or self._output_file or str(runtime.output_file if runtime else "")
        if not resolved_output:
            return ResultSummary()

        path = Path(resolved_output)
        if not path.exists():
            return ResultSummary(output_file=str(path), exists=False, error="Output file does not exist.")

        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            return ResultSummary(output_file=str(path), exists=True, error=str(exc))
        try:
            return ResultSummary(
                output_file=str(path),
                exists=True,
                term_library_count=_count_sheet_rows(workbook, TERM_LIBRARY_SHEET),
                review_count=_count_sheet_rows(workbook, REVIEW_SHEET),
                failure_count=_count_sheet_rows(workbook, FAILURE_SHEET),
                nontrans_regex_count=_count_sheet_rows(workbook, NONTRANS_REGEX_SHEET),
            )
        finally:
            workbook.close()

    def _run_task(self, settings: AppSettings, task_input: Optional[TaskInput], resume: bool) -> None:
        service = self.service_factory(
            settings=settings,
            runtime_store=self.runtime_store,
            paths=self.paths,
            logger=self.logger,
            log_callback=self._append_log,
            progress_callback=self._record_progress,
            stop_requested=self._stop_event.is_set,
        )
        try:
            output_file = asyncio.run(service.run(task_input, resume=resume))
        except TaskCancelledError as exc:
            with self._lock:
                self._last_error = str(exc)
                self._logs.append(str(exc))
        except Exception as exc:
            self.logger.exception("\u4efb\u52a1\u6267\u884c\u5931\u8d25")
            with self._lock:
                self._last_error = str(exc)
                self._logs.append(str(exc))
                _track_text_preprocess_finish(self._active_telemetry_context, success=False)
        else:
            with self._lock:
                self._output_file = output_file
                _track_text_preprocess_finish(self._active_telemetry_context, success=True)
        finally:
            with self._lock:
                self._is_running = False

    def _append_log(self, message: str) -> None:
        with self._lock:
            self._logs.append(str(message))

    def _record_progress(self, payload: Dict[str, object]) -> None:
        with self._lock:
            self._last_progress = dict(payload or {})


def runtime_to_snapshot(runtime: Optional[RuntimeTaskState]) -> TaskSnapshot:
    if runtime is None:
        return TaskSnapshot(is_running=False)
    stats = dict(runtime.stats or {})
    return TaskSnapshot(
        is_running=False,
        can_resume=True,
        stage=runtime.stage or "IDLE",
        progress_current=int(stats.get("progress_current", 0) or 0),
        progress_total=int(stats.get("progress_total", 0) or 0),
        current_batch=int(stats.get("current_batch", 0) or 0),
        total_batches=int(stats.get("total_batches", 0) or 0),
        success_count=int(stats.get("success_count", 0) or 0),
        failure_count=int(stats.get("failure_count", 0) or 0),
        retry_count=int(stats.get("retry_count", 0) or 0),
        current_concurrency=int(stats.get("current_concurrency", 0) or 0),
        output_file=runtime.output_file,
        last_error=runtime.last_error,
        stats=stats,
    )


def _count_sheet_rows(workbook, sheet_name: str) -> int:
    if sheet_name not in workbook.sheetnames:
        return 0
    sheet = workbook[sheet_name]
    count = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if any(cell not in (None, "") for cell in row):
            count += 1
    return count


def _build_telemetry_context(
    *,
    settings: AppSettings,
    task_input: Optional[TaskInput],
    runtime: Optional[RuntimeTaskState],
) -> Dict[str, object]:
    input_config = task_input.to_dict() if task_input is not None else dict(runtime.input_config if runtime else {})
    extraction_mode = str(input_config.get("extraction_mode", "terms") or "terms").strip()
    provider_name = str(settings.provider_name or "DeepSeek")
    provider = settings.provider_settings.get(provider_name)
    model_name = str(provider.model if provider is not None else "")
    nontrans_stage = dict(settings.input_defaults.get("nontrans_stage_settings", {}) or {})
    recall_stage = dict(settings.input_defaults.get("term_recall_stage_settings", {}) or {})
    review_stage = dict(settings.input_defaults.get("term_review_stage_settings", {}) or {})

    if extraction_mode == "terms":
        uses_ai = True
        thinking_enabled = bool(
            (bool(nontrans_stage.get("ai_discovery_enabled", True)) or bool(nontrans_stage.get("ai_regex_generation_enabled", True)))
            and bool(nontrans_stage.get("enable_thinking", False))
        ) or bool(recall_stage.get("enable_thinking", False)) or bool(review_stage.get("enable_thinking", False))
    else:
        uses_ai = bool(nontrans_stage.get("ai_discovery_enabled", True)) or bool(
            nontrans_stage.get("ai_regex_generation_enabled", True)
        )
        thinking_enabled = uses_ai and bool(nontrans_stage.get("enable_thinking", False))

    return {
        "extraction_mode": extraction_mode,
        "uses_ai": uses_ai,
        "thinking_enabled": thinking_enabled,
        "model_name": model_name,
    }


def _track_text_preprocess_start(context: Dict[str, object]) -> None:
    track_event("task_start.text_preprocess")
    if str(context.get("extraction_mode")) == "nontrans_only":
        track_event("task_mode.nontrans_only")
    else:
        track_event("task_mode.term_extract")
    if not bool(context.get("uses_ai")):
        return
    track_event("task_start.ai_tool")
    if bool(context.get("thinking_enabled")):
        track_event("model_mode.thinking_enabled")
    tier = infer_model_tier(str(context.get("model_name", "")))
    if tier == "flash":
        track_event("model_tier.flash")
    elif tier == "pro":
        track_event("model_tier.pro")


def _track_text_preprocess_finish(context: Dict[str, object], *, success: bool) -> None:
    if success:
        track_event("task_success.text_preprocess")
    else:
        track_event("task_fail.text_preprocess")
    if not bool(context.get("uses_ai")):
        return
    if success:
        track_event("task_success.ai_tool")
    else:
        track_event("task_fail.ai_tool")
