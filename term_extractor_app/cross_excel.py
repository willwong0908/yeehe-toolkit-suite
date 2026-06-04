"""Cross-Excel search and merge helpers."""

from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Sequence

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .core import _open_excel_file
from .storage import get_app_paths


SUPPORTED_EXTENSIONS = {".xlsx", ".xls", ".xlsm"}
MERGE_RESULT_PREFIXES = ("合并结果_", "merge_result_")


@dataclass
class CrossExcelSearchMatch:
    file_name: str
    sheet_name: str
    row_index: int
    row_values: List[str]
    matched_columns: List[int]

    def to_dict(self) -> Dict[str, object]:
        return {
            "file_name": self.file_name,
            "sheet_name": self.sheet_name,
            "row_index": self.row_index,
            "row_values": list(self.row_values),
            "matched_columns": list(self.matched_columns),
        }


def _is_valid_excel_file(path: Path) -> bool:
    if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
        return False
    if path.name.startswith("~$"):
        return False
    lowered_name = path.name.lower()
    if any(lowered_name.startswith(prefix.lower()) for prefix in MERGE_RESULT_PREFIXES):
        return False
    return True


def get_all_excel_files(folder_path: str) -> List[Path]:
    root = Path(folder_path)
    if not root.exists() or not root.is_dir():
        raise ValueError("文件夹路径无效。")
    files = [path for path in root.glob("**/*") if path.is_file() and _is_valid_excel_file(path)]
    return sorted(files, key=lambda item: str(item).lower())


def collect_all_headers(excel_files: Sequence[Path]) -> tuple[List[str], Dict[str, Dict[str, List[str]]]]:
    all_headers = set()
    file_sheet_headers: Dict[str, Dict[str, List[str]]] = {}

    for file_path in excel_files:
        with _open_excel_file(str(file_path)) as excel_file:
            file_sheet_headers[str(file_path)] = {}
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(
                    excel_file,
                    sheet_name=sheet_name,
                    nrows=0,
                    keep_default_na=False,
                )
                headers = [str(column) for column in df.columns]
                file_sheet_headers[str(file_path)][sheet_name] = headers
                all_headers.update(headers)

    return sorted(all_headers), file_sheet_headers


def scan_cross_excel_folder(folder_path: str) -> Dict[str, object]:
    excel_files = get_all_excel_files(folder_path)
    headers, file_sheet_headers = collect_all_headers(excel_files)
    return {
        "folder_path": folder_path,
        "file_count": len(excel_files),
        "headers": headers,
        "files": [path.name for path in excel_files],
        "file_sheet_headers": file_sheet_headers,
    }


def _trim_row_values(values: Iterable[object]) -> List[str]:
    normalized = ["" if value is None else str(value) for value in values]
    last_non_empty = -1
    for index, value in enumerate(normalized):
        if str(value).strip():
            last_non_empty = index
    if last_non_empty < 0:
        return []
    return normalized[: last_non_empty + 1]


def search_excel_rows(folder_path: str, query: str, limit: int = 300) -> Dict[str, object]:
    search_text = str(query or "").strip()
    if not search_text:
        raise ValueError("请输入要搜索的内容。")

    excel_files = get_all_excel_files(folder_path)
    lowered = search_text.casefold()
    matches: List[CrossExcelSearchMatch] = []
    scanned_rows = 0

    for file_path in excel_files:
        with _open_excel_file(str(file_path)) as excel_file:
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(
                    excel_file,
                    sheet_name=sheet_name,
                    keep_default_na=False,
                    header=None,
                )
                for row_offset, row_values in enumerate(df.itertuples(index=False, name=None), start=1):
                    scanned_rows += 1
                    trimmed = _trim_row_values(row_values)
                    if not trimmed:
                        continue
                    matched_columns = [
                        index
                        for index, cell_value in enumerate(trimmed)
                        if lowered in str(cell_value or "").casefold()
                    ]
                    if not matched_columns:
                        continue
                    matches.append(
                        CrossExcelSearchMatch(
                            file_name=file_path.name,
                            sheet_name=sheet_name,
                            row_index=row_offset,
                            row_values=trimmed,
                            matched_columns=matched_columns,
                        )
                    )
                    if len(matches) >= max(1, int(limit or 300)):
                        return {
                            "query": search_text,
                            "file_count": len(excel_files),
                            "scanned_rows": scanned_rows,
                            "truncated": True,
                            "items": [item.to_dict() for item in matches],
                        }

    return {
        "query": search_text,
        "file_count": len(excel_files),
        "scanned_rows": scanned_rows,
        "truncated": False,
        "items": [item.to_dict() for item in matches],
    }


def copy_cell_style(source_cell, target_cell) -> None:
    if source_cell is None or not source_cell.has_style:
        return
    if source_cell.fill:
        target_cell.fill = copy(source_cell.fill)
    if source_cell.font:
        target_cell.font = copy(source_cell.font)
    if source_cell.alignment:
        target_cell.alignment = copy(source_cell.alignment)
    if source_cell.border:
        target_cell.border = copy(source_cell.border)
    if source_cell.number_format:
        target_cell.number_format = source_cell.number_format


def _cross_excel_output_dir() -> Path:
    output_dir = get_app_paths().output_dir / "cross_excel_merge"
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir


def merge_excel_files_by_headers(
    folder_path: str,
    selected_headers: Sequence[str],
    apply_format: bool = True,
) -> Dict[str, object]:
    headers = [str(item).strip() for item in selected_headers if str(item).strip()]
    if not headers:
        raise ValueError("请至少选择一个表头。")

    excel_files = get_all_excel_files(folder_path)
    _, file_sheet_headers = collect_all_headers(excel_files)
    merged_data: List[pd.DataFrame] = []
    merged_formats: List[Dict[str, object]] = []

    for file_path in excel_files:
        file_key = str(file_path)
        for sheet_name, sheet_headers in (file_sheet_headers.get(file_key) or {}).items():
            matching_headers = [header for header in headers if header in sheet_headers]
            if not matching_headers:
                continue

            with _open_excel_file(str(file_path)) as excel_file:
                df = pd.read_excel(
                    excel_file,
                    sheet_name=sheet_name,
                    keep_default_na=False,
                )
                df.columns = [str(column) for column in df.columns]

            columns_to_keep = [column for column in headers if column in df.columns]
            if not columns_to_keep:
                continue

            filtered = df[columns_to_keep].copy()
            filtered.insert(0, "来源文件", file_path.name)
            filtered.insert(1, "来源Sheet", sheet_name)
            merged_data.append(filtered)

            if not apply_format:
                for _ in range(len(filtered)):
                    merged_formats.append({column: None for column in filtered.columns})
                continue

            try:
                workbook = load_workbook(file_path, data_only=False)
                worksheet = workbook[sheet_name]
                header_row = 1
                header_map: Dict[str, int] = {}
                for column_index in range(1, worksheet.max_column + 1):
                    header_value = worksheet.cell(row=header_row, column=column_index).value
                    normalized = str(header_value).strip() if header_value is not None else ""
                    for column_name in columns_to_keep:
                        if column_name.strip() == normalized:
                            header_map[column_name] = column_index
                            break
                for data_row_index in range(2, len(df) + 2):
                    row_format: Dict[str, object] = {"来源文件": None, "来源Sheet": None}
                    for column_name in columns_to_keep:
                        source_column = header_map.get(column_name)
                        row_format[column_name] = (
                            worksheet.cell(row=data_row_index, column=source_column)
                            if source_column
                            else None
                        )
                    merged_formats.append(row_format)
                workbook.close()
            except Exception:
                for _ in range(len(filtered)):
                    merged_formats.append({column: None for column in filtered.columns})

    if not merged_data:
        raise ValueError("没有找到包含所选表头的数据。")

    result_df = pd.concat(merged_data, ignore_index=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    output_path = _cross_excel_output_dir() / f"合并结果_{timestamp}.xlsx"

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        result_df.to_excel(writer, sheet_name="合并数据", index=False)

    workbook = load_workbook(output_path)
    worksheet = workbook["合并数据"]

    if apply_format:
        for row_index, row_format in enumerate(merged_formats, start=2):
            for column_index, column_name in enumerate(result_df.columns, start=1):
                target_cell = worksheet.cell(row=row_index, column=column_index)
                copy_cell_style(row_format.get(column_name), target_cell)

    for column_index, _ in enumerate(result_df.columns, start=1):
        worksheet.column_dimensions[get_column_letter(column_index)].width = 15

    workbook.save(output_path)
    workbook.close()

    return {
        "output_file": str(output_path),
        "row_count": int(len(result_df)),
        "column_count": int(len(result_df.columns)),
        "selected_headers": headers,
        "apply_format": bool(apply_format),
        "output_dir": str(output_path.parent),
    }
