"""Excel diff helpers for the Yeehe toolkit."""

from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Callable, Dict, Iterable, Iterator, List, Sequence, Tuple
from uuid import uuid4

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from .storage import get_app_paths


SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm"}
DIFF_RESULT_PREFIX = "excel_diff_result_"
DIFF_PREVIEW_LIMIT = 1000


@dataclass(frozen=True)
class DiffRecord:
    filename_a: str
    filename_b: str
    sheet: str
    cell_address: str
    value_a: str
    value_b: str
    file_path_a: str
    file_path_b: str

    @property
    def pair_label(self) -> str:
        if self.filename_a == self.filename_b:
            return self.filename_a
        return f"{self.filename_a} <> {self.filename_b}"

    @property
    def search_blob(self) -> str:
        return " ".join(
            [
                self.filename_a,
                self.filename_b,
                self.sheet,
                self.cell_address,
                self.value_a,
                self.value_b,
            ]
        ).lower()

    def to_dict(self) -> Dict[str, str]:
        return {
            "filename_a": self.filename_a,
            "filename_b": self.filename_b,
            "sheet": self.sheet,
            "cell_address": self.cell_address,
            "value_a": self.value_a,
            "value_b": self.value_b,
            "file_path_a": self.file_path_a,
            "file_path_b": self.file_path_b,
        }


@dataclass(frozen=True)
class CompareMeta:
    mode_label: str
    files_in_a: int
    files_in_b: int
    matched_pairs: int
    diff_count: int

    def to_dict(self) -> Dict[str, object]:
        return {
            "mode_label": self.mode_label,
            "files_in_a": self.files_in_a,
            "files_in_b": self.files_in_b,
            "matched_pairs": self.matched_pairs,
            "diff_count": self.diff_count,
        }


def normalize_path(value: str) -> str:
    text = str(value or "").strip().strip("\"'")
    if not text:
        return ""
    return str(Path(text).expanduser())


def format_cell_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value != value:
        return ""
    return str(value)


def compare_text_values(
    val_a: object,
    val_b: object,
    *,
    ignore_case: bool = False,
    trim_whitespace: bool = True,
) -> bool:
    text_a = format_cell_value(val_a)
    text_b = format_cell_value(val_b)
    if trim_whitespace:
        text_a = text_a.strip()
        text_b = text_b.strip()
    if ignore_case:
        text_a = text_a.casefold()
        text_b = text_b.casefold()
    return text_a != text_b


def _is_supported_excel_file(path: Path) -> bool:
    if not path.is_file():
        return False
    if path.name.startswith("~$"):
        return False
    return path.suffix.lower() in SUPPORTED_EXTENSIONS


def scan_excel_files(path: str) -> List[str]:
    normalized = normalize_path(path)
    if not normalized:
        return []
    root = Path(normalized)
    if not root.exists():
        return []
    if _is_supported_excel_file(root):
        return [str(root.resolve())]
    if not root.is_dir():
        return []
    files = [item.resolve() for item in root.rglob("*") if _is_supported_excel_file(item)]
    return sorted(str(item) for item in files)


def match_excel_files(files_a: Sequence[str], files_b: Sequence[str]) -> List[Tuple[str, str, str]]:
    index_a: Dict[str, str] = {}
    index_b: Dict[str, str] = {}

    for file_path in files_a:
        index_a.setdefault(Path(file_path).name, file_path)
    for file_path in files_b:
        index_b.setdefault(Path(file_path).name, file_path)

    common_names = sorted(set(index_a) & set(index_b))
    return [(index_a[name], index_b[name], name) for name in common_names]


def compare_excel_files(
    file_a: str,
    file_b: str,
    *,
    ignore_case: bool = False,
    trim_whitespace: bool = True,
) -> List[DiffRecord]:
    diffs: List[DiffRecord] = []
    workbook_a = load_workbook(file_a, data_only=True)
    workbook_b = load_workbook(file_b, data_only=True)

    try:
        common_sheets = sorted(set(workbook_a.sheetnames) & set(workbook_b.sheetnames))
        filename_a = Path(file_a).name
        filename_b = Path(file_b).name

        for sheet_name in common_sheets:
            sheet_a = workbook_a[sheet_name]
            sheet_b = workbook_b[sheet_name]
            max_row = max(sheet_a.max_row, sheet_b.max_row)
            max_col = max(sheet_a.max_column, sheet_b.max_column)

            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    value_a = sheet_a.cell(row, col).value
                    value_b = sheet_b.cell(row, col).value
                    if not compare_text_values(
                        value_a,
                        value_b,
                        ignore_case=ignore_case,
                        trim_whitespace=trim_whitespace,
                    ):
                        continue

                    diffs.append(
                        DiffRecord(
                            filename_a=filename_a,
                            filename_b=filename_b,
                            sheet=sheet_name,
                            cell_address=f"{get_column_letter(col)}{row}",
                            value_a=format_cell_value(value_a),
                            value_b=format_cell_value(value_b),
                            file_path_a=str(Path(file_a).resolve()),
                            file_path_b=str(Path(file_b).resolve()),
                        )
                    )
    finally:
        workbook_a.close()
        workbook_b.close()

    return diffs


def iter_compare_excel_records(
    file_a: str,
    file_b: str,
    *,
    ignore_case: bool = False,
    trim_whitespace: bool = True,
) -> Iterator[DiffRecord]:
    workbook_a = load_workbook(file_a, data_only=True)
    workbook_b = load_workbook(file_b, data_only=True)

    try:
        common_sheets = sorted(set(workbook_a.sheetnames) & set(workbook_b.sheetnames))
        filename_a = Path(file_a).name
        filename_b = Path(file_b).name

        for sheet_name in common_sheets:
            sheet_a = workbook_a[sheet_name]
            sheet_b = workbook_b[sheet_name]
            max_row = max(sheet_a.max_row, sheet_b.max_row)
            max_col = max(sheet_a.max_column, sheet_b.max_column)

            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    value_a = sheet_a.cell(row, col).value
                    value_b = sheet_b.cell(row, col).value
                    if not compare_text_values(
                        value_a,
                        value_b,
                        ignore_case=ignore_case,
                        trim_whitespace=trim_whitespace,
                    ):
                        continue
                    yield DiffRecord(
                        filename_a=filename_a,
                        filename_b=filename_b,
                        sheet=sheet_name,
                        cell_address=f"{get_column_letter(col)}{row}",
                        value_a=format_cell_value(value_a),
                        value_b=format_cell_value(value_b),
                        file_path_a=str(Path(file_a).resolve()),
                        file_path_b=str(Path(file_b).resolve()),
                    )
    finally:
        workbook_a.close()
        workbook_b.close()


def compare_paths(
    path_a: str,
    path_b: str,
    *,
    ignore_case: bool = False,
    trim_whitespace: bool = True,
    progress_callback: Callable[[str], None] | None = None,
) -> tuple[List[DiffRecord], CompareMeta]:
    normalized_a = normalize_path(path_a)
    normalized_b = normalize_path(path_b)
    progress = progress_callback or (lambda _message: None)

    file_a = Path(normalized_a)
    file_b = Path(normalized_b)
    is_file_a = _is_supported_excel_file(file_a)
    is_file_b = _is_supported_excel_file(file_b)

    if is_file_a and is_file_b:
        progress("正在比较两个 Excel 文件")
        diffs = compare_excel_files(
            str(file_a),
            str(file_b),
            ignore_case=ignore_case,
            trim_whitespace=trim_whitespace,
        )
        return diffs, CompareMeta("文件对文件", 1, 1, 1, len(diffs))

    progress("正在扫描目录中的 Excel 文件")
    files_a = scan_excel_files(normalized_a)
    files_b = scan_excel_files(normalized_b)
    pairs = match_excel_files(files_a, files_b)

    all_diffs: List[DiffRecord] = []
    total = len(pairs)
    for index, (paired_a, paired_b, name) in enumerate(pairs, start=1):
        progress(f"正在比较 {name} ({index}/{total})")
        all_diffs.extend(
            compare_excel_files(
                paired_a,
                paired_b,
                ignore_case=ignore_case,
                trim_whitespace=trim_whitespace,
            )
        )

    mode_label = "目录对目录" if file_a.is_dir() or file_b.is_dir() else "混合模式"
    meta = CompareMeta(
        mode_label=mode_label,
        files_in_a=len(files_a),
        files_in_b=len(files_b),
        matched_pairs=len(pairs),
        diff_count=len(all_diffs),
    )
    return all_diffs, meta


def excel_color_from_hex(color_hex: str) -> str:
    cleaned = str(color_hex or "").replace("#", "").strip()
    if len(cleaned) == 6:
        return f"FF{cleaned.upper()}"
    if len(cleaned) == 8:
        return cleaned.upper()
    raise ValueError(f"无法识别的颜色值: {color_hex}")


def apply_highlight_to_records(records: Sequence[DiffRecord], target: str, color_hex: str) -> tuple[int, int]:
    if target not in {"A", "B"}:
        raise ValueError("target 必须是 A 或 B")

    grouped: Dict[str, Dict[str, set[str]]] = {}
    for record in records:
        workbook_path = record.file_path_a if target == "A" else record.file_path_b
        sheet_map = grouped.setdefault(workbook_path, {})
        sheet_map.setdefault(record.sheet, set()).add(record.cell_address)

    fill_color = excel_color_from_hex(color_hex)
    fill = PatternFill(fill_type="solid", start_color=fill_color, end_color=fill_color)
    changed_cells = 0

    for workbook_path, sheet_map in grouped.items():
        workbook = load_workbook(workbook_path)
        try:
            for sheet_name, cell_addresses in sheet_map.items():
                if sheet_name not in workbook.sheetnames:
                    continue
                worksheet = workbook[sheet_name]
                for address in sorted(cell_addresses):
                    worksheet[address].fill = fill
                    changed_cells += 1
            workbook.save(workbook_path)
        finally:
            workbook.close()

    return changed_cells, len(grouped)


def _diff_output_dir() -> Path:
    output_dir = get_app_paths().output_dir / "excel_diff"
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir


def _diff_cache_dir() -> Path:
    cache_dir = _diff_output_dir() / "cache"
    cache_dir.mkdir(parents=True, exist_ok=True)
    return cache_dir


def run_compare_to_cache(
    path_a: str,
    path_b: str,
    *,
    ignore_case: bool = False,
    trim_whitespace: bool = True,
    progress_callback: Callable[[str], None] | None = None,
    preview_limit: int = DIFF_PREVIEW_LIMIT,
) -> Dict[str, object]:
    normalized_a = normalize_path(path_a)
    normalized_b = normalize_path(path_b)
    progress = progress_callback or (lambda _message: None)
    file_a = Path(normalized_a)
    file_b = Path(normalized_b)
    is_file_a = _is_supported_excel_file(file_a)
    is_file_b = _is_supported_excel_file(file_b)

    result_id = uuid4().hex
    cache_path = _diff_cache_dir() / f"{result_id}.jsonl"
    preview_records: List[Dict[str, str]] = []
    diff_count = 0

    if is_file_a and is_file_b:
        mode_label = "文件对文件"
        files_in_a = 1
        files_in_b = 1
        matched_pairs = 1
        pairs = [(str(file_a), str(file_b), file_a.name)]
        progress("正在比较两个 Excel 文件")
    else:
        mode_label = "目录对目录" if file_a.is_dir() or file_b.is_dir() else "混合模式"
        files_a = scan_excel_files(normalized_a)
        files_b = scan_excel_files(normalized_b)
        pairs = match_excel_files(files_a, files_b)
        files_in_a = len(files_a)
        files_in_b = len(files_b)
        matched_pairs = len(pairs)
        progress("正在扫描目录中的 Excel 文件")

    with cache_path.open("w", encoding="utf-8") as handle:
        total_pairs = len(pairs)
        for index, (paired_a, paired_b, name) in enumerate(pairs, start=1):
            if not (is_file_a and is_file_b):
                progress(f"正在比较 {name} ({index}/{total_pairs})")
            for record in iter_compare_excel_records(
                paired_a,
                paired_b,
                ignore_case=ignore_case,
                trim_whitespace=trim_whitespace,
            ):
                item = record.to_dict()
                handle.write(json.dumps(item, ensure_ascii=False) + "\n")
                diff_count += 1
                if len(preview_records) < max(1, int(preview_limit or DIFF_PREVIEW_LIMIT)):
                    preview_records.append(item)

    meta = CompareMeta(
        mode_label=mode_label,
        files_in_a=files_in_a,
        files_in_b=files_in_b,
        matched_pairs=matched_pairs,
        diff_count=diff_count,
    )
    return {
        "result_id": result_id,
        "cache_file": str(cache_path),
        "preview_records": preview_records,
        "preview_limit": max(1, int(preview_limit or DIFF_PREVIEW_LIMIT)),
        "preview_truncated": diff_count > len(preview_records),
        "meta": meta.to_dict(),
        "total_count": diff_count,
    }


def iter_cached_diff_records(cache_file: str, *, query: str = "") -> Iterator[DiffRecord]:
    lowered = str(query or "").strip().lower()
    path = Path(cache_file)
    if not path.exists():
        return
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            text = str(line or "").strip()
            if not text:
                continue
            item = json.loads(text)
            record = DiffRecord(
                filename_a=str(item.get("filename_a", "") or ""),
                filename_b=str(item.get("filename_b", "") or ""),
                sheet=str(item.get("sheet", "") or ""),
                cell_address=str(item.get("cell_address", "") or ""),
                value_a=str(item.get("value_a", "") or ""),
                value_b=str(item.get("value_b", "") or ""),
                file_path_a=str(item.get("file_path_a", "") or ""),
                file_path_b=str(item.get("file_path_b", "") or ""),
            )
            if lowered and lowered not in record.search_blob:
                continue
            yield record


def read_cached_diff_preview(
    cache_file: str,
    *,
    query: str = "",
    limit: int = DIFF_PREVIEW_LIMIT,
) -> Dict[str, object]:
    preview: List[Dict[str, str]] = []
    matched_count = 0
    safe_limit = max(1, int(limit or DIFF_PREVIEW_LIMIT))
    for record in iter_cached_diff_records(cache_file, query=query):
        matched_count += 1
        if len(preview) < safe_limit:
            preview.append(record.to_dict())
    return {
        "records": preview,
        "matched_count": matched_count,
        "preview_limit": safe_limit,
        "preview_truncated": matched_count > len(preview),
    }


def export_diff_records(records: Iterable[DiffRecord], output_path: str = "") -> Dict[str, object]:
    destination = str(output_path or "").strip()
    if destination:
        final_path = Path(destination)
        if final_path.suffix.lower() not in SUPPORTED_EXTENSIONS:
            final_path = final_path.with_suffix(".xlsx")
        final_path.parent.mkdir(parents=True, exist_ok=True)
    else:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        final_path = _diff_output_dir() / f"{DIFF_RESULT_PREFIX}{stamp}.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "差异列表"
    headers = ["文件A", "文件B", "Sheet", "单元格", "A内容", "B内容"]
    worksheet.append(headers)
    diff_count = 0
    max_lengths = [len(item) for item in headers]
    for row in records:
        worksheet.append(
            [
                row.filename_a,
                row.filename_b,
                row.sheet,
                row.cell_address,
                row.value_a,
                row.value_b,
            ]
        )
        diff_count += 1
        values = [row.filename_a, row.filename_b, row.sheet, row.cell_address, row.value_a, row.value_b]
        for index, value in enumerate(values):
            max_lengths[index] = max(max_lengths[index], min(len(str(value or "")), 80))

    for column_index, max_len in enumerate(max_lengths, start=1):
        worksheet.column_dimensions[get_column_letter(column_index)].width = max(12, min(max_len + 2, 100))

    workbook.save(final_path)
    workbook.close()

    return {
        "output_file": str(final_path),
        "output_dir": str(final_path.parent),
        "diff_count": diff_count,
    }


def export_cached_diff_records(cache_file: str, output_path: str = "", *, query: str = "") -> Dict[str, object]:
    return export_diff_records(iter_cached_diff_records(cache_file, query=query), output_path)


def apply_highlight_from_cache(cache_file: str, target: str, color_hex: str, *, query: str = "") -> tuple[int, int]:
    records = list(iter_cached_diff_records(cache_file, query=query))
    return apply_highlight_to_records(records, target, color_hex)
